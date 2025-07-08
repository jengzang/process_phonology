"""
运行这个函数，即可根据“例字”文件里的字生成声母/韵母
更改声母/韵母输出在第45行
"""
import tkinter as tk
from tkinter import filedialog
import pandas as pd
from sort_characters import processing_examples_vowels, processing_examples_consonants
from openpyxl import load_workbook
from openpyxl.comments import Comment


def extract_rhyme_from_files() -> None:
    """
    弹出文件选择对话框，允许用户选择多个文件，提取韵母并保存结果。
    """
    # 创建文件选择对话框
    root = tk.Tk()
    root.withdraw()  # 不显示主窗口
    file_paths = filedialog.askopenfilenames(title="选择文件", filetypes=[("TSV Files", "*.tsv")])  # 只选择tsv文件

    if not file_paths:
        print("没有选择任何文件！")
        return

    # 设置固定的元音文件路径
    vowel_file_path = r"C:\Users\joengzaang\myfiles\杂文件\声韵处理\例字.xlsx"

    # 读取 '例字.xlsx' 文件，保留例字列
    example_df = pd.read_excel(vowel_file_path)

    # 创建一个新的 DataFrame 用来存储所有的韵母结果
    combined_results = example_df[["例字"]].copy()

    # 保存每个文件的韵母和批注
    file_results_dict = {}

    for file_path in file_paths:
        try:
            print(f"正在处理文件: {file_path}")

            # 读取 .tsv 文件
            tsv_df = pd.read_csv(file_path, sep='\t', encoding='utf-8')

            ###############################韵母##########################
            file_results = processing_examples_vowels(tsv_df, vowel_file_path)
            ###############################韵母##########################

            ###############################声母##########################
            # file_results = processing_examples_consonants(tsv_df, vowel_file_path)
            ###############################声母##########################

            # 将该文件的韵母列添加到 combined_results 中
            file_name = file_path.split("/")[-1].replace(".tsv", "")
            combined_results[file_name] = file_results['声韵']
            file_results_dict[file_name] = file_results['批注']  # 保存批注内容

        except Exception as e:
            print(f"处理文件 {file_path} 时出错: {e}")

    # 生成最终输出的文件
    output_file = r"C:\Users\joengzaang\myfiles\杂文件\声韵处理\聲韻表\聲韻表_新生成.xlsx"
    combined_results.to_excel(output_file, index=False, engine='openpyxl')
    print(f"所有声韵提取结果已保存到：{output_file}")

    # 插入批注
    wb = load_workbook(output_file)
    ws = wb.active

    # 假设韵母在 DataFrame 中的列，从第二行开始插入批注
    for file_name, comments in file_results_dict.items():
        column_index = combined_results.columns.get_loc(file_name) + 1  # 获取韵母列的索引（+1 是因为 openpyxl 是 1 基础的）

        # 遍历每个韵母单元格，插入批注
        for i, comment in enumerate(comments, start=2):  # start=2 是因为 Excel 数据从第二行开始
            # 去除批注内容的空格，并检查是否为空
            if pd.notna(comment) and comment.strip():  # 如果批注内容非空（去除空格后）
                cell = ws.cell(row=i, column=column_index)
                cell.comment = Comment(comment, "System")

    # 保存 Excel 文件
    wb.save(output_file)
    print(f"批注已成功插入并保存到：{output_file}")


if __name__ == "__main__":
    extract_rhyme_from_files()
