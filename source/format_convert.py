#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
本腳本將整合音典、跳跳老鼠、縣志三種格式的完整字表提取邏輯，
支援 .tsv、.xlsx、.xls、.docx 格式
根據預設表或用戶選擇對應格式，轉換為 #漢字 音標 解釋 的 .tsv 文件。
"""

import csv
import os
import re
from itertools import product

import docx
import pandas as pd
from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from opencc import OpenCC
from openpyxl import load_workbook
from xlrd import open_workbook

from source.config import ZHENGZI_PATH, MULCODECHAR_PATH

opencc_s2t = OpenCC('s2t.json')


# def get_tsv_name(path):
#     return os.path.splitext(path)[0] + ".tsv"

def get_tsv_name(xls):
    name = os.path.basename(xls)
    name = re.sub(r" ?(\(\d{0,3}\))+$", "", name.rsplit(".", 1)[0]) + ".tsv"
    return os.path.join(os.path.dirname(__file__), name)


def xls_to_tsv(xls_path, page=0):
    def is_xls(fname):
        return fname.endswith("xls") or fname.endswith("xlsx")

    def process_fs(v):
        t = type(v)
        if t is float or t is int:
            return "%d" % v
        if v is None:
            return ""
        return str(v).strip().replace("\t", " ").replace("\n", " ")

    def process_xlsx_fs(v):
        t = type(v)
        if t is float or t is int:
            return "%d" % v
        if v is None:
            return ""
        if t is str:
            return v.strip().replace("\t", " ").replace("\n", " ")
        cells = []
        for i in v:
            if isinstance(i, str):
                cells.append(i.strip())
                continue
            if isinstance(i, (int, float)):
                cells.append("%d" % i)
                continue
            text = i.text
            tag = ""
            if i.font.underline == "single":
                tag = "-"
            elif i.font.underline == "double":
                tag = "="
            if tag:
                text = "".join([j + tag for j in text])
            if i.font.vertAlign == "subscript" or (i.font.size and i.font.size < 10.0):
                text = f"({text})"
            cells.append(text)
        return "".join(cells).replace(")(", "").strip()

    def get_tsv_name(path):
        return os.path.splitext(path)[0] + ".tsv"

    print(f"[INFO] Starting conversion: {xls_path}")
    if not os.path.exists(xls_path):
        print(f"[ERROR] File does not exist: {xls_path}")
        return

    tsv_path = get_tsv_name(xls_path)
    print(f"[INFO] Target TSV path: {tsv_path}")

    lines = []
    header_written = False
    num_columns = 0

    if xls_path.endswith(".xlsx"):
        print("[INFO] Detected .xlsx file")
        wb = load_workbook(xls_path, data_only=True, rich_text=True)
        sheet = wb.worksheets[page]
        print(f"[INFO] Loaded worksheet: {sheet.title}")
        for row_idx, row in enumerate(sheet.rows):
            cols = [process_xlsx_fs(cell.value) for cell in row[:50]]
            if any(cols):
                if not header_written:
                    num_columns = len(cols)
                    header_written = True
                cols += [""] * (num_columns - len(cols))
                lines.append("\t".join(cols[:num_columns]) + "\n")
    else:
        print("[INFO] Detected .xls file")
        wb = open_workbook(xls_path)
        sheet = wb.sheet_by_index(page)
        print(f"[INFO] Loaded sheet: {sheet.name}")
        for i in range(sheet.nrows):
            row = sheet.row_values(i)
            cols = [process_fs(cell) for cell in row]
            if any(cols):
                if not header_written:
                    num_columns = len(cols)
                    header_written = True
                cols += [""] * (num_columns - len(cols))
                lines.append("\t".join(cols[:num_columns]) + "\n")

    print(f"[INFO] Writing {len(lines)} rows to TSV")
    with open(tsv_path, "w", encoding="utf-8", newline="\n") as f:
        f.writelines(lines)

    print(f"[INFO] Conversion complete: {tsv_path}")
    return tsv_path


def run2text(run):
    if isinstance(run, docx.text.hyperlink.Hyperlink):
        return "".join(map(run2text, run.runs))
    tag = ""
    if run.font.underline == docx.enum.text.WD_UNDERLINE.SINGLE:
        tag = "-"
    elif run.font.underline == docx.enum.text.WD_UNDERLINE.DOUBLE:
        tag = "="
    elif run.font.underline == docx.enum.text.WD_UNDERLINE.WAVY:
        tag = chr(0x1AB6)
    elif run._r.xpath("*/w:em[@w:val='dot']"):
        tag = chr(0x0323)
    text = run.text
    if tag:
        text = "".join([i + tag for i in text])
    if run.font.subscript or (run.font.size and run.font.size < docx.shared.Pt(9)):
        if text.startswith("{") and text.endswith("}"):
            pass
        elif text.startswith("[") and text.endswith("]"):
            pass
        else:
            text = f"{{{text}}}"
    return text


def docx_to_tsv(doc):
    tsv = get_tsv_name(doc)
    if not os.path.exists(doc): return
    if os.path.exists(tsv):
        xtime = os.path.getmtime(doc)
        ttime = os.path.getmtime(tsv)
        if ttime >= xtime: return
    lines = []
    Doc = Document(doc)
    for each in Doc._body._element:
        if isinstance(each, docx.oxml.table.CT_Tbl):
            t = Table(each, Doc)
            for row in t.rows:
                行 = ""
                cells = row.cells
                for i, cell in enumerate(cells):
                    if cell in cells[:i]: continue
                    for p in cell.paragraphs:
                        行 += "".join(map(run2text, p.iter_inner_content())).replace("\t", "").replace("\n", "")
                    行 += "\t"
                lines.append(
                    行.replace("}~", "~}").replace("~{", "{~").replace("}{", "").replace("[}", "}[").replace("{h}",
                                                                                                             "h").strip())
        elif isinstance(each, docx.oxml.text.paragraph.CT_P):
            element = Paragraph(each, Doc)
            行 = "".join(map(run2text, element.iter_inner_content())).replace("}~", "~}").replace("~{", "{~").replace(
                "}{", "").replace("[}", "}[").replace("{h}", "h")
            lines.append(行)
    行 = "\n".join(lines).replace("}\n{", "").replace("\n}", "}\n")
    with open(tsv, "w", encoding="utf-8", newline="\n") as t:
        t.write(行)
    return tsv


def convert_to_tsv_if_needed(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext in [".xlsx", ".xls"]:
        # wb = load_workbook(filepath, data_only=True)
        # sheet = wb.active
        # lines = []
        # for row in sheet.iter_rows(values_only=True):
        #     if all(cell is None for cell in row): continue
        #     line = "\t".join([str(cell).strip() if cell is not None else "" for cell in row])
        #     lines.append(line + "\n")
        # tsv_path = get_tsv_name(filepath)
        # with open(tsv_path, "w", encoding="utf-8") as f:
        #     f.writelines(lines)
        # return tsv_path
        return xls_to_tsv(filepath)
    elif ext == ".docx":
        return docx_to_tsv(filepath)
    else:
        return filepath


# ========== 繁體轉換函數 ==========
def s2t_pro(字組, level=1):
    variant_file = os.path.join(os.path.dirname(__file__), ZHENGZI_PATH)
    mulcode_file = os.path.join(os.path.dirname(__file__), MULCODECHAR_PATH)

    normVariants = {}
    stVariants = {}
    n2o_dict = {}

    # 讀取正字表
    for 行 in open(variant_file, encoding="utf-8"):
        if 行.startswith("#"):
            continue  # 行首為註解，跳過

        行 = 行.rstrip("\n")
        列 = 行.split("\t")
        if len(列) < 2:
            continue

        原字 = 列[0].strip()
        對應字串 = 列[1].split("#")[0].strip()  # 去除 # 後的註解
        候選字列表 = 對應字串.split()

        if level == 1:
            if "#" in 行:
                continue  # 含 # 的行不處理
            if len(候選字列表) > 1:
                continue  # 多候選字，不處理

        stVariants[原字] = 對應字串

    # 讀取 mulcodechar.dt
    for 行 in open(mulcode_file, encoding="utf-8"):
        if not 行 or 行[0] == "#":
            continue
        列 = 行.strip().split("-")
        if len(列) < 2:
            continue
        n2o_dict[列[0]] = 列[1]

    def n2o(s):
        return ''.join(n2o_dict.get(i, i) for i in s)

    result_chars = []
    mapping = []

    for 字 in 字組:
        原字 = 字
        對應字串 = stVariants.get(字, None)

        if 對應字串 is None and level == 2:
            對應字串 = opencc_s2t.convert(原字)
            # print(f"【OpenCC】{原字} → {對應字串}")  # Debug 用
        elif 對應字串 is None:
            對應字串 = 原字

        對應字串 = n2o(對應字串)

        # 保留候選字列表
        if " " in 對應字串:
            候選 = 對應字串.split()
        else:
            候選 = [對應字串]
        mapping.append((原字, 候選))
        result_chars.extend(候選)

    clean_str = ''.join(result_chars)

    return clean_str, mapping


# ========== 音典格式處理 ==========
def process_音典(file, level=1, output_path=None):
    print(f"[開始] 處理檔案：{file}")

    # cc = OpenCC('s2t' if level == 1 else 't2s')
    # def s2t(text):
    #     return cc.convert(text)

    file = convert_to_tsv_if_needed(file)
    print(f"[轉換] 轉為 TSV 路徑：{file}")

    rows = []
    simplified_rows = []

    with open(file, encoding="utf-8") as f:
        lines = [line.rstrip("\n").split("\t") for line in f if line.strip() and not line.startswith("#")]

    if not lines:
        print("⚠️ 無有效數據，檔案內容為空或格式錯誤")
        os.makedirs("data", exist_ok=True)
        with open("logs/error.txt", "a", encoding="utf-8") as f:
            f.write(f"⚠️ [{file}] 無有效數據，檔案內容為空或格式錯誤\t【format_convert->process_音典】\n")
        return

    header = lines[0]
    print(f"[分析] 表頭：{header}")

    col_map = {
        '漢字': ['漢字_程序改名', '單字', '单字', '漢字', 'phrase'],
        '音標': ['IPA_程序改名', 'IPA', 'ipa', '音標', 'syllable'],
        '解釋': ['注釋_程序改名', '注释', '注釋', '解釋', 'notes']
    }
    index = {}
    for std_key, aliases in col_map.items():
        for i, name in enumerate(header):
            if name.strip().lower() in [a.lower() for a in aliases]:
                index[std_key] = i
                print(f"✅ 欄位對應：{std_key} → 第 {i + 1} 欄（{name}）")
                break

    if '漢字' not in index or '音標' not in index:
        print("❌ 欄位對應失敗，請確認有『漢字』與『音標』欄位")
        os.makedirs("data", exist_ok=True)
        with open("logs/error.txt", "a", encoding="utf-8") as f:
            f.write(f"❌ [{file}]欄位對應失敗，請確認有『漢字』與『音標』欄位\t【format_convert->process_音典】\n")
        return

    print(f"[處理] 開始掃描資料行，共 {len(lines) - 1} 筆")

    delimiters = [';', '；', '/', '、', ',', '，']

    def split_field(field):
        for delim in delimiters:
            field = field.replace(delim, '∥')  # 統一分隔符為 ∥
        return [f.strip() for f in field.split('∥') if f.strip()]

    def get_field(parts, field_name):
        idx = index.get(field_name)
        if idx is not None and idx < len(parts):
            return parts[idx].strip()
        return ""

    def process_pair(word, phon, note, row_num):
        clean_str, mapping = s2t_pro(word, level)
        mapping = dict(mapping)
        phon_units = phon.strip().split()
        word_len_match = len(word) == len(phon_units)

        if word_len_match:
            for ch, p in zip(word, phon_units):
                candidates = mapping.get(ch, [ch])
                for cand in candidates:
                    rows.append([cand, p, note])
                    if cand != ch:
                        simplified_rows.append([cand, p, note, "簡"])
                        print(f"[簡體一對多] 第 {row_num} 行：{ch} → {cand}")
        else:
            rows.append([clean_str, phon, note])
            if clean_str != word:
                simplified_rows.append([clean_str, phon, note, "簡"])
                print(f"[fallback] 第 {row_num} 行：{word} → {clean_str}")

    print(f"[處理] 開始掃描資料行，共 {len(lines) - 1} 筆")

    for row_num, parts in enumerate(lines[1:], start=2):
        word_raw = get_field(parts, '漢字')
        phon_raw = get_field(parts, '音標')
        note = get_field(parts, '解釋')

        if not word_raw or not phon_raw:
            continue

        word_list = split_field(word_raw)
        phon_list = split_field(phon_raw)

        if not word_list or not phon_list:
            print(f"⚠️ 跳過第 {row_num} 行，因為漢字或音標清單為空")
            continue

        if len(word_list) > 1 and len(phon_list) > 1:
            # ✅ 無論等長與否，始終做笛卡爾積
            print(f"[笛卡爾積] 第 {row_num} 行：{word_list} × {phon_list}")
            for word, phon in product(word_list, phon_list):
                process_pair(word, phon, note, row_num)

        elif len(word_list) > 1 and len(phon_list) == 1:
            # ✅ 多對一
            print(f"[多對一] 第 {row_num} 行：{word_list} × {phon_list[0]}")
            for word in word_list:
                process_pair(word, phon_list[0], note, row_num)

        elif len(word_list) == 1 and len(phon_list) > 1:
            # ✅ 一對多
            print(f"[一對多] 第 {row_num} 行：{word_list[0]} × {phon_list}")
            for phon in phon_list:
                process_pair(word_list[0], phon, note, row_num)

        else:
            # fallback 合併處理
            word = ''.join(word_list)
            phon = ' '.join(phon_list)
            # print(f"[fallback] 第 {row_num} 行：{word} → {phon}")
            process_pair(word, phon, note, row_num)

    outpath = output_path or (os.path.splitext(file)[0] + ".tsv")
    print(f"[輸出] 寫入主檔案：{outpath}")

    with open(outpath, "w", encoding="utf-8", newline="\n") as out:
        writer = csv.writer(out, delimiter="\t")
        writer.writerow(["#漢字", "音標", "解釋"])
        writer.writerows(rows)

    # simp_path = os.path.splitext(file)[0] + ".簡.tsv"
    # if simplified_rows:
    #     print(f"[簡體] 共發現 {len(simplified_rows)} 筆簡體詞彙，寫入：{simp_path}")
    #     with open(simp_path, "w", encoding="utf-8", newline="\n") as out:
    #         writer = csv.writer(out, delimiter="\t")
    #         writer.writerow(["#漢字", "音標", "解釋", "繁簡"])
    #         writer.writerows(simplified_rows)

    print(f"✅ 全部處理完成：{outpath}")


# ========== 跳跳老鼠格式處理 ==========
def process_跳跳老鼠(file, level=1, output_path=None):
    print(f"📄 開始處理文件：{file}")
    rows = []
    simplified_rows = []

    # 選擇繁→簡或簡→繁
    # converter = OpenCC('s2t' if level == 1 else 't2s')
    # def s2t(text):
    #     return converter.convert(text)

    # 讀取 Excel（僅第一張表）
    wb = load_workbook(file, data_only=True)
    sheet = wb.active

    def parse_row(line, line_num):
        parts = [str(c).strip() if c is not None else "" for c in line]
        if len(parts) < 2:
            print(f"⚠️ 第 {line_num} 行欄位不足，跳過：{parts}")
            return []
        phon = parts[0]
        組 = parts[1]
        if not phon or not 組:
            print(f"⚠️ 第 {line_num} 行缺音或字，跳過")
            return []
        result = []
        matches = re.findall(r"(.)(?:\{(.*?)\}|\[(.*?)\])?", 組)
        print(f"🔍 第 {line_num} 行組拆分：{matches}")
        for 字, 註1, 註2 in matches:
            註 = 註1 or 註2 or ""
            print(f"🧩 字：{字}，音：{phon}，註：{註}")
            result.append((字, phon, 註))
        return result

    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not row or str(row[0]).startswith("#"):
            continue
        parsed = parse_row(row, i)
        for 字, 音, 註 in parsed:
            clean_str, mapping = s2t_pro(字, level)
            mapping = dict(mapping)
            candidates = mapping.get(字, [字])  # 支援多候選

            for cand in candidates:
                rows.append([cand, 音, 註])
                if cand != 字:
                    simplified_rows.append([cand, 音, 註, "簡"])
                    print(f"🔁 字形轉換：{字} → {cand}")

    outpath = output_path or os.path.splitext(file)[0] + ".tsv"
    with open(outpath, "w", encoding="utf-8", newline="\n") as out:
        writer = csv.writer(out, delimiter="\t")
        writer.writerow(["#漢字", "音標", "解釋"])
        writer.writerows(rows)
    print(f"✅ 主檔輸出完成：{outpath}")

    simp_path = os.path.splitext(file)[0] + ".簡.tsv"
    if simplified_rows:
        with open(simp_path, "w", encoding="utf-8", newline="\n") as out:
            writer = csv.writer(out, delimiter="\t")
            writer.writerow(["#漢字", "音標", "解釋", "繁簡"])
            writer.writerows(simplified_rows)
        print(f"[簡體] 共發現 {len(simplified_rows)} 筆簡體詞彙，寫入：{simp_path}")

    print(f"🎉 全部處理完成，共 {len(rows)} 條記錄")
    # print(f"✅ 輸出：{outpath}")


# ========== 縣志格式處理 ==========
def process_縣志(file, level=1, output_path=None):
    cc = OpenCC('s2t')
    rows = []
    simplified_rows = []
    debug = True

    # def s2t(text, level=1):
    #     return cc.convert(text)

    def process_lines(行):
        行 = 行.strip()
        if not 行:
            return None
        if 行.startswith("#"):
            return 行
        行 = re.sub(r":\[", "\t[", 行)
        行 = 行.replace("(", "{").replace(")", "}")
        行 = re.sub(r"\[(\d+)\]", r"［\1］", 行)
        行 = re.sub(r"［([^\d]+.*?)］", r"[\1]", 行)
        return 行

    ext = os.path.splitext(file)[1].lower()
    if ext in [".xlsx", ".xls"]:
        df = pd.read_excel(file, sheet_name=0, header=None)
        lines = [
            "\t".join([str(cell) for cell in row if pd.notna(cell)]).strip()
            for _, row in df.iterrows()
        ]
        print(f"📖 讀取 Excel：{file}")
    else:
        encodings = ["utf-8", "utf-8-sig", "big5", "gb18030"]
        for enc in encodings:
            try:
                with open(file, encoding=enc) as f:
                    lines = f.readlines()
                print(f"📖 使用編碼：{enc}")
                break
            except UnicodeDecodeError:
                continue
        else:
            raise UnicodeDecodeError("❌ 無法讀取文件，請確認編碼格式")

    total, skipped, simplified_count = 0, 0, 0

    for lineno, line in enumerate(lines, 1):
        total += 1
        raw_line = line
        line = process_lines(line)
        if line is None:
            skipped += 1
            continue

        if line.startswith("#漢字"):
            skipped += 1
            continue
        if line.startswith("#"):
            continue

        parts = line.split("\t")
        if len(parts) < 2:
            if debug:
                print(f"⚠️ 跳過行 {lineno}（分欄不足）: {raw_line.strip()}")
            skipped += 1
            continue

        拼音 = parts[0].strip()
        for cell in parts[1:]:
            matches = re.findall(r"[［\[](\d+[a-z]?)[］\]](.+?)(?=([［\[]\d|$))", cell)
            if not matches:
                if debug:
                    print(f"⚠️ 無音節匹配 行 {lineno}: {cell}")
                continue

            for 調號, 義項, _ in matches:
                if debug:
                    print(f"🔎 行 {lineno}：拼音={拼音}, 調號={調號}, 義項={義項}")

                # 逐字掃描義項，若某字後緊跟註釋，就綁定在那個字上
                i = 0
                while i < len(義項):
                    字 = 義項[i]
                    註 = ""
                    if i + 1 < len(義項) and 義項[i + 1] in "{｛":
                        m = re.match(r"[{｛]([^{}｛｝]+)[}｝]", 義項[i + 1:])
                        if m:
                            註 = m.group(1)
                            i += len(m.group(0))  # 跳過整個 {註釋}
                    i += 1
                    字 = 字.strip()
                    if not 字:
                        if debug:
                            print(f"⚠️ 空白字 行 {lineno} 義項：{義項}")
                        continue
                    clean_str, mapping = s2t_pro(字, level)
                    mapping = dict(mapping)
                    candidates = mapping.get(字, [字])  # 支援多候選繁體字

                    音標 = f"{拼音}{調號}"
                    for cand in candidates:
                        row = [cand, 音標, 註]
                        rows.append(row)
                        if cand != 字:
                            simplified_rows.append(row + ["簡"])
                            simplified_count += 1
                            if debug:
                                print(f"🔁 字形轉換：{字} → {cand}")

    outpath = output_path or os.path.splitext(file)[0] + ".tsv"
    with open(outpath, "w", encoding="utf-8", newline="\n") as out:
        writer = csv.writer(out, delimiter="\t")
        writer.writerow(["#漢字", "音標", "解釋"])
        writer.writerows(rows)
    print(f"✅ 主檔輸出完成：{outpath}")

    simp_path = os.path.splitext(file)[0] + ".簡.tsv"
    if simplified_rows:
        with open(simp_path, "w", encoding="utf-8", newline="\n") as out:
            writer = csv.writer(out, delimiter="\t")
            writer.writerow(["#漢字", "音標", "解釋", "繁簡"])
            writer.writerows(simplified_rows)
        print(f"[簡體] 共發現 {len(simplified_rows)} 筆簡體詞彙，寫入：{simp_path}")

    print(f"📊 行數統計：總行數 {total}, 跳過 {skipped} 行, 標註簡體 {simplified_count} 條")


# test = ['台','高']
# result = s2t_pro(test,level=2)
# print(result)
