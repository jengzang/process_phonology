"""
运行这个函数，即可根据“聲韻”文件里的层级结构按照中古音地位整理所有字
更改声母/韵母/声调输出在第69、73、77行
在第304行选择使用“聲韻”文件中的列名，可以自己根据中古音制作层级结构，用“-”隔开即可
在第305行输入要处理的分区，不同分区用空格隔开，输入“全部”则都处理。
第17行可以添加模糊音，把对应声韵放在一行内输出
第203行设置当声韵少于一定比例时，不独立一行输出
"""

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.comments import Comment
from gets import get_consonants_from_tsv, get_vowels_from_tsv, get_tones_from_tsv
from collections import defaultdict

from matching import choose_tsv_files

# 聲韻模糊映射：鍵為原聲韻，值為歸類用的主聲韻
MERGE_MAP = {}

MERGE_MAP.update({k: "kʷ" for k in ["kw", "kᵘ", "kᵛ", "kʋ", "kʷ", "kv"]})
MERGE_MAP.update({k: "kʰw" for k in ["kʷʰ", "kʰʷ", "kʰᵘ", "kʰᵛ", "kʰʋ", "kʋʰ", "kʰw", "kvʰ"]})
MERGE_MAP.update({k: "pʰʋ" for k in ["pʰw", "pʰᵘ", "pʰʋ"]})
MERGE_MAP.update({k: "tʰw" for k in ["tʰᵘ", "tʰw", "tʰʋ"]})
MERGE_MAP.update({k: "ʔ" for k in ["(ʔ)", "∅", "ʔ", "ˀ"]})
MERGE_MAP.update({k: "ʋ" for k in ["v", "ʋ", "vʋ", "w"]})
MERGE_MAP.update({k: "h" for k in ["h", "ɦ", "ɦʰ", "xʱ", "hɦ", "hʱ", "ʰ"]})
MERGE_MAP.update({k: "hʷ" for k in ["hʷ", "hw", "hʋ", "ɦʋ"]})
MERGE_MAP.update({k: "x" for k in ["x", "xʱ", "xɣ", "ɣ", "χ"]})
MERGE_MAP.update({k: "xʷ" for k in ["xv", "xʋ", "xʷ", "xᵊ", "xᶷ"]})
MERGE_MAP.update({k: "d" for k in ["d", "d̥", "ɗ", "ɗw"]})
MERGE_MAP.update({k: "dz" for k in ["dz", "d̥z̥"]})
MERGE_MAP.update({k: "dʑ" for k in ["dʑ", "d̥ʑ̥"]})
MERGE_MAP.update({k: "fw" for k in ["fʋ", "fw", "fv", "fʰ", "fʱ", "f", "̊f"]})
MERGE_MAP.update({k: "l" for k in ["l", "l̥", "l̩"]})
MERGE_MAP.update({k: "m" for k in ["m", "m̥", "m̩", "m͡b"]})
MERGE_MAP.update({k: "mʷ" for k in ["mʷ", "mw", "mʋ"]})
MERGE_MAP.update({k: "mʰ" for k in ["mʰ", "mɦ", "mʱ"]})
MERGE_MAP.update({k: "sʷ" for k in ["sw", "sʋ", "sʷ"]})
MERGE_MAP.update({k: "tʰ" for k in ["tʰʰ", "tʱ", "tʰ"]})
MERGE_MAP.update({k: "ŋʷ" for k in ["ŋʷ", "ŋw", "ŋʋ"]})
MERGE_MAP.update({k: "ŋ" for k in ["ŋ", "ŋ̊", "ŋɡ", "ŋ͡ɡ", "ng", "nɡ"]})
MERGE_MAP.update({k: "ɡ" for k in ["ɡ", "g", "ɡ̊", "ᵑɡ"]})
MERGE_MAP.update({k: "b" for k in ["b̥", "ɓw", "ɓ", "ᵐb", "b", "bv"]})


def extract_levels(level_df, category_column):
    level_dict = defaultdict(list)
    max_level = 0
    for _, row in level_df.iterrows():
        levels = str(row[category_column]).split("-")
        max_level = max(max_level, len(levels))
        level_dict[tuple(levels)].append(row["單字"])
    return level_dict, max_level


def build_flat_levels(level_dict, max_level):
    expanded = []
    for levels, chars in level_dict.items():
        for char in chars:
            padded = list(levels) + [None] * (max_level - len(levels))
            expanded.append(padded + [char])
    return pd.DataFrame(expanded, columns=[f"level{i + 1}" for i in range(max_level)] + ["char"])


def collect_consonants(tsv_path):
    ########################聲母#################################
    # df = get_consonants_from_tsv(tsv_path, char_list="all")
    ########################聲母#################################

    ########################韻母#################################
    df = get_vowels_from_tsv(tsv_path, char_list="all")
    ########################韻母#################################

    ########################聲調#################################
    # df = get_tones_from_tsv(tsv_path, char_list="all")
    ########################聲調#################################

    phonetic_map = defaultdict(list)
    for _, row in df.iterrows():
        phonetic_map[row["汉字"]].append((row["声韵"], row["音标"]))
    return phonetic_map


def process(tsv_paths, excel_path, category_column):
    # print(tsv_paths)
    raw_df = pd.read_excel(excel_path, sheet_name="層級")
    level_df = raw_df[['單字', category_column]].dropna()
    print(f"级别数据框创建完成，去除缺失值后共有 {len(level_df)} 行。")
    level_dict, max_level = extract_levels(level_df, category_column)
    flat_df = build_flat_levels(level_dict, max_level)

    output_dir = r"C:\Users\joengzaang\myfiles\杂文件\声韵处理\arrange"
    os.makedirs(output_dir, exist_ok=True)
    save_path = os.path.join(output_dir, f"{category_column}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    # 允許 tsv_paths 中包含 "_" 作為空欄佔位符
    real_paths = [p for p in tsv_paths if p != "_"]
    placeholder_names = ["_"] * tsv_paths.count("_")
    real_names = [os.path.splitext(os.path.basename(p))[0] for p in real_paths]
    tsv_names = []

    # 重建 tsv_names，對應原始順序（保留 "_" 的位置）
    for p in tsv_paths:
        if p == "_":
            tsv_names.append("_")
        else:
            tsv_names.append(os.path.splitext(os.path.basename(p))[0])

    print(f"识别到的TSV文件（含佔位）: {tsv_names}")

    # 只為非 "_" 的檔案建立聲韻對應表
    phonetic_maps = {
        name: collect_consonants(path)
        for name, path in zip(tsv_names, tsv_paths)
        if name != "_"
    }

    for level_idx in range(1, max_level + 1):
        sheet = wb.create_sheet(title=f"第{level_idx}級")
        level_cols = [f"level{i}" for i in range(1, level_idx + 1)]
        header = level_cols[:]
        for name in tsv_names:
            if name == "_":
                header += ["", ""]  # 保留空欄佔位
            else:
                header += [f"{name}_聲韻", f"{name}_轄字"]
                # header += [f"{name}", f"{name}_聲韻", f"{name}_轄字"]
        sheet.append(header)

        subset = flat_df.dropna(subset=[f"level{level_idx}"])
        level_groups = defaultdict(list)
        for _, row in subset.iterrows():
            key = tuple(row[c] for c in level_cols)
            level_groups[key].append(row["char"])
        print(f"第 {level_idx} 级的分组完成，共有 {len(level_groups)} 个唯一组。")

        for level_key, chars in level_groups.items():
            if not tsv_names:
                sheet.append(list(level_key) + ["", ""])
                continue
            all_consonants = set()
            tsv_consonant_map = {}
            merged_chars_by_file = {}
            merged_class_map = defaultdict(list)

            # Step 0: 每个文件内的有效字数（出现在 phonetic_maps 中）
            total_chars_per_file = {}
            for name in tsv_names:
                if name == "_":
                    continue
                total_chars_per_file[name] = sum(1 for char in chars if char in phonetic_maps[name])

            # Step 1: 遍历所有 tsv，按子類聲韻统计数据，同时建立主類→子類映射
            for name in tsv_names:
                if name == "_":
                    continue  # 跳過佔位符，不查表
                cmap = defaultdict(list)
                for char in chars:
                    if char in phonetic_maps[name]:
                        for cons, _ in phonetic_maps[name][char]:
                            main_cons = MERGE_MAP.get(cons, cons)
                            cmap[cons].append(char)
                            all_consonants.add(main_cons)
                            if cons not in merged_class_map[main_cons]:
                                merged_class_map[main_cons].append(cons)
                tsv_consonant_map[name] = cmap

            # Step 2: 按主類聲韻排序（字典序）
            # 自定义排序列表
            custom_order = [
                'p', 'pʰ', 't', 'tʰ', 'k', 'kʰ', 'f', 'ʋ', 'ɸ', 'h',
                'x', 'l', 'n', 'm', 'ŋ', 'ɲ', 'ȵ', 'j', 'z', 's', 'ʃ',
                'ʂ', 'ɕ', 'θ', 'ɬ', 'b', 'd', 'g', 'ʒ', 'ʑ', 'ʐ'
                'ʦ', 'ʧ', 'ʨ', 'tʂ', 'tɹ', 'tr', 'tθ', 'dz', 'dʑ', 'dʐ', 'dʒ'
                'ʦʰ', 'ʧʰ', 'ʨʰ', 'tʂʰ', 'tɹʰ', 'trʰ', 'tθʰ', 'dzʰ', 'dʑʰ', 'dʐʰ', 'dʒʰ'
                'ʔ', 'a', 'ia', 'ua', 'ᴀ', 'ɑ', 'æ', 'ɐ', 'iɐ', 'uɐ',
                'ə', 'iə', 'uə', 'ᴇ', 'ɛ', 'œ', 'iɛ', 'uɛ', 'ɜ', 'ɞ', 'ʌ',
                'ɔ', 'iɔ', 'uɔ', 'o', 'io', 'uo', 'ɤ', 'ɵ', 'ɘ',
                'ø', 'iø', 'e', 'ie', 'ʊ', 'u', 'ɯ', 'y', 'i', 'ɿ', 'ʮ',
                '陰平', '陰平甲', '陰平乙', '陽平', '陽平甲', '陽平乙', '陰上', '陰上甲', '陰上乙',
                '陽上', '陽上甲', '陽上乙', '陰去', '陰去甲', '陰去乙', '陽去', '陽去甲', '陽去乙',
                '陰入', '上陰入', '下陰入', '陽入', '上陽入', '下陽入', '變調', '變調1', '變調2', '輕聲',
            ]

            # 按照 custom_order 排序 all_consonants
            sorted_consonants = sorted(all_consonants,
                                       key=lambda cons: [
                                           custom_order.index(cons[i:i + 2]) if cons[i:i + 2] in custom_order else
                                           custom_order.index(cons[i]) if cons[i] in custom_order else float('inf')
                                           for i in range(len(cons))
                                       ])

            # print(sorted_consonants)
            # sorted_consonants = sorted(all_consonants)
            small_class_cache = []

            # Step 3: 逐主類聲韻生成一行数据
            for cons in sorted_consonants:
                row = list(level_key)
                merged_chars_by_file = {}  # 必须每轮清零
                small_class_files = set()  # 标记哪些文件要合并
                for name in tsv_names:
                    if name == "_":
                        continue
                    cmap = tsv_consonant_map.get(name, {})
                    merged_cons_list = merged_class_map.get(cons, [cons])
                    actual_cons_present = [c for c in merged_cons_list if c in cmap]

                    if not actual_cons_present:
                        continue

                    merged_chars = []
                    for c in actual_cons_present:
                        merged_chars.extend(cmap[c])
                    merged_chars = list(dict.fromkeys(merged_chars))  # 保序去重

                    total = total_chars_per_file.get(name, 1)  # 避免除以 0
                    proportion = len(merged_chars) / total
                    # 小于0.07的声韵不独立显示
                    if proportion < 0.07:
                        small_class_files.add(name)

                    merged_chars_by_file[name] = merged_chars

                # 判断是否整行合并或部分写入
                if small_class_files:
                    trimmed_merged_map = {
                        name: merged_chars_by_file[name]
                        for name in small_class_files
                        if name in merged_chars_by_file
                    }
                    small_class_cache.append((cons, trimmed_merged_map))

                # ✅ 写入主行（部分文件写入）
                row_data = []
                for name in tsv_names:
                    if name == "_":
                        row_data += ["", ""]
                    elif name in merged_chars_by_file and name not in small_class_files:
                        chars = merged_chars_by_file[name]
                        row_data += [cons, "".join(chars)]
                    else:
                        row_data += ["", ""]
                # ✅ 如果整行都是空的（該主類所有文件都為小主類），跳過
                if all(cell == "" for cell in row_data):
                    continue
                row += row_data
                sheet.append(row)

                # Step 4: 添加批註
                col_base = len(level_cols)
                for i, name in enumerate(tsv_names):
                    if name == "_" or name in small_class_files:
                        continue
                    if name in merged_chars_by_file:
                        for char in merged_chars_by_file[name]:
                            phonetics = phonetic_maps[name].get(char, [])
                            if len(phonetics) > 1:
                                readings = ", ".join(p[1] for p in phonetics)
                                new_text = f"{char}：{readings}"
                                target_col = col_base + i * 2 + 2
                                cell = sheet.cell(row=sheet.max_row, column=target_col)
                                if cell.comment:
                                    cell.comment = Comment(cell.comment.text + f"\n{new_text}", "不羈")
                                else:
                                    cell.comment = Comment(new_text, "不羈")

            # Step 5: 合併寫入所有小占比主類（若有）
            if small_class_cache:
                row = list(level_key)
                row_data = []

                for name in tsv_names:
                    if name == "_":
                        row_data += ["", ""]
                        continue

                    cons_label_lines = []
                    cons_char_lines = []

                    for cons, merged_map in small_class_cache:
                        if name in merged_map:
                            cons_label_lines.append(cons)
                            cons_char_lines.append("".join(merged_map[name]))

                    if cons_label_lines:
                        row_data.append("\n".join(cons_label_lines))
                        row_data.append("\n".join(cons_char_lines))
                    else:
                        row_data += ["", ""]

                row += row_data
                sheet.append(row)

                # ✅ 批註
                col_base = len(level_cols)
                for i, name in enumerate(tsv_names):
                    if name == "_":
                        continue
                    for cons, merged_map in small_class_cache:
                        if name in merged_map:
                            chars = merged_map[name]
                            for char in chars:
                                phonetics = phonetic_maps[name].get(char, [])
                                if len(phonetics) > 1:
                                    readings = ", ".join(p[1] for p in phonetics)
                                    new_text = f"{char}：{readings}"
                                    target_col = col_base + i * 2 + 2
                                    cell = sheet.cell(row=sheet.max_row, column=target_col)
                                    if cell.comment:
                                        cell.comment = Comment(cell.comment.text + f"\n{new_text}", "不羈")
                                    else:
                                        cell.comment = Comment(new_text, "不羈")
    wb.save(save_path)
    print(f"✅ 已導出：{save_path}")


# -------------------------------
if __name__ == "__main__":
    EXCEL_PATH = r"C:\Users\joengzaang\myfiles\杂文件\声韵处理\聲韻.xlsx"
    CATEGORY_COLUMN = "韻母簡"  # 改完记得改上面调用的函数
    tsv_files, *_ = choose_tsv_files("嶺南 嶺西 廣中")  # 嶺南 嶺西 廣中 嶺東 閩 湘贛 浙南 兩浙

    if tsv_files:
        process(tsv_files, EXCEL_PATH, CATEGORY_COLUMN)
    else:
        print("⚠️ 未選擇任何TSV文件。")
