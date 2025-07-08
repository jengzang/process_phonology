"""
运行这个函数，即可根据“统计”文件计算声韵频率
需要做的是:把声韵、例字复制到统计里（第一行是xx_聲韻 xx_轄字)
然后在该程序第26行设置要跳过的列数（即第一个地名出现之前有多少列）
第77行选择排序的依据（根据“聲韻”里的“順序”工作表的列名进行排序）
"""

import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.comments import Comment
from hanziconv import HanziConv

from matching import process_and_sort_locations

# === 1. 檔案路徑 ===
input_path = r"C:\Users\joengzaang\myfiles\杂文件\声韵处理\统计.xlsx"
output_path = r"C:\Users\joengzaang\myfiles\杂文件\声韵处理\声韵分析\聲韻頻率統計.xlsx"
# abbreviation_path = r"C:\Users\joengzaang\myfiles\杂文件\声韵处理\漢字音典字表檔案（長期更新）.csv"
aspiration_path = r"C:\Users\joengzaang\myfiles\杂文件\声韵处理\聲韻.xlsx"

# === 2. 讀取資料 ===
df = pd.read_excel(input_path)
# 忽略多少列在这里更改，例如忽略3列下面就填3
df_data = df.iloc[:, 1:]

# === 3. 地點順序 ===
original_locations = []
for i in range(0, len(df_data.columns), 2):
    col_name = df_data.columns[i]
    if isinstance(col_name, str) and col_name.endswith('聲韻'):
        loc_name = col_name.rsplit('_', 1)[0]
        original_locations.append(loc_name)

# === 4. 地點與分區排序處理 ===
locations, partitions, matched_locations, unmatched_locations, partition_map = process_and_sort_locations(original_locations)


# === 5. 統計頻率 ===
location_freqs = {}
location_comments = {}
overall_counter = defaultdict(int)
overall_total = 0

for loc in locations:
    if loc == "__SEP__":
        continue
    rhyme_col = f"{loc}_聲韻"
    char_col = f"{loc}_轄字"
    if rhyme_col in df_data.columns and char_col in df_data.columns:
        sub_df = df_data[[rhyme_col, char_col]].dropna()
        rhyme_counts = defaultdict(int)
        rhyme_chars = defaultdict(str)
        total_chars = 0

        for _, row in sub_df.iterrows():
            rhyme = row[rhyme_col]
            chars = str(row[char_col])
            count = len(chars)
            total_chars += count
            rhyme_counts[rhyme] += count
            rhyme_chars[rhyme] += chars
            overall_counter[rhyme] += count
            overall_total += count

        freq_percent = {k: round(v / total_chars * 100, 1) for k, v in rhyme_counts.items()}
        location_freqs[loc] = freq_percent
        location_comments[loc] = rhyme_chars


# === 6. 整理預計表 ===
all_rhymes = sorted(set(r for freqs in location_freqs.values() for r in freqs))
# print(all_rhymes)

# 按順序排序 all_rhymes
aspiration_df = pd.read_excel(aspiration_path, sheet_name="順序")
rhyme_order = aspiration_df["送氣"].dropna().tolist()

# 保證順序唯一並保留原順序
seen = set()
rhyme_order_unique = [r for r in rhyme_order if not (r in seen or seen.add(r))]

all_rhymes = [r for r in rhyme_order_unique if r in all_rhymes] + [r for r in all_rhymes if r not in rhyme_order_unique]
print("最終聲韻排序：", all_rhymes)


freq_table = pd.DataFrame(index=all_rhymes)

cols = {}
for loc in locations:
    if loc == "__SEP__":
        cols[loc] = pd.Series([''] * len(all_rhymes), index=all_rhymes)
    else:
        cols[loc] = pd.Series([location_freqs.get(loc, {}).get(r, 0.0) for r in all_rhymes], index=all_rhymes)
freq_table = pd.concat(cols, axis=1)

# === 7. 總頻率 ===
overall_freq = {k: round(v / overall_total * 100, 1) for k, v in overall_counter.items()}
freq_table['總頻率'] = [overall_freq.get(r, 0.0) for r in all_rhymes]
# freq_table.sort_values(by='總頻率', ascending=False, inplace=True)

# === 8. 格式化 ===
freq_table_percent = freq_table.apply(lambda col: col.map(lambda x: f"{x:.1f}%" if isinstance(x, float) and x > 0 else ""))

# === 9. 寫入 Excel 並加入分區 ===
wb = Workbook()
ws = wb.active
ws.title = "聲韻頻率"

headers = ["聲韻"] + locations + ["總頻率"]
region_row = ["分區"] + partitions + [""]

ws.append(region_row)
ws.append(headers)

for r_idx, (rhyme, row) in enumerate(freq_table_percent.iterrows(), start=3):
    ws.cell(row=r_idx, column=1, value=rhyme)
    for c_idx, loc in enumerate(locations, start=2):
        val = row[loc] if loc in row else ""
        ws.cell(row=r_idx, column=c_idx, value=val)
    ws.cell(row=r_idx, column=1 + len(locations) + 1, value=row['總頻率'])

# === 10. 加入批評 ===
for r_idx, rhyme in enumerate(freq_table_percent.index, start=3):
    for c_idx, loc in enumerate(locations, start=2):
        if loc == "__SEP__":
            continue
        cell = ws.cell(row=r_idx, column=c_idx)
        comment_text = location_comments.get(loc, {}).get(rhyme, "")
        if comment_text:
            cell.comment = Comment(comment_text, "系統")

wb.save(output_path)
print(f"已輸出至：{output_path}")

