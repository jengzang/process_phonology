import os
import re
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl import load_workbook


# 文件选择窗口
def select_files():
    root = Tk()
    root.withdraw()  # 隐藏主窗口
    file_paths = filedialog.askopenfilenames(title='选择Excel文件', filetypes=[('Excel Files', '*.xlsx')])
    return list(file_paths)


# 获取工作表名中的所有汉字
def extract_all_chinese_from_sheetname(sheetname):
    # 使用正则表达式提取工作表名中的所有汉字字符
    chinese_part = ''.join(re.findall(r'[\u4e00-\u9fff]+', sheetname))
    if chinese_part:
        return chinese_part
    return 'processed_file'


# 查找可能的列名，如果不存在原始列名，寻找拼音或英文版列名
def find_column(df, primary_name, alternative_names):
    # 检查是否存在主列名
    if primary_name in df.columns:
        return primary_name
    # 检查是否存在拼音或英文版列名
    for alt_name in alternative_names:
        if alt_name in df.columns:
            return alt_name
    # 若没有找到任何匹配列，返回None并输出警告
    print(f"警告: 未找到列 {primary_name} 或其替代列 {alternative_names}")
    return None


# 合并"聲母"、"韻母"、"調值"的函数，处理空值的情况
def merge_syllable(row, shengmu_col, yunmu_col, diaozhi_col):
# def merge_syllable(row, shengmu_col, yunmu_col):
    shengmu = row[shengmu_col] if pd.notna(row[shengmu_col]) else ""
    yunmu = row[yunmu_col] if pd.notna(row[yunmu_col]) else ""
    diaozhi = str(int(row[diaozhi_col])) if pd.notna(row[diaozhi_col]) else ""
    return shengmu + yunmu + diaozhi
    # return shengmu + yunmu


# 清理備註列，删除"文讀"，并确保空值处理
def clean_notes(notes):
    if pd.isna(notes):
        return ""  # 如果是NaN，则返回空字符串
    cleaned = notes.replace('文讀', '')  # 删除"文讀"
    return cleaned if cleaned.strip() else ""  # 如果删除后为空，保持空


# 处理单个Excel文件
def process_file(file_path):
    print(f"正在处理文件: {file_path}")  # 输出正在处理的文件

    # 使用openpyxl引擎打开工作簿
    workbook = load_workbook(file_path, read_only=True)
    sheetnames = workbook.sheetnames  # 获取所有工作表名称
    first_sheet_name = sheetnames[0]  # 假设处理第一个工作表

    # 使用pandas读取第一个工作表
    df = pd.read_excel(file_path, sheet_name=first_sheet_name, engine='openpyxl')

    # 查找“聲母”、“韻母”、“調值”、“字”、“備註”或它们的拼音或英文版本
    shengmu_col = find_column(df, '聲母', ['ShengMu'])
    yunmu_col = find_column(df, '韻母', ['YunMu'])
    diaozhi_col = find_column(df, '調值', ['DiaoZhi'])
    char_col = find_column(df, '字', ['Char'])
    comment_col = find_column(df, '備註', ['Comment'])

    # 如果缺少“聲母”、“韻母”或“調值”列，则跳过该文件
    # if not shengmu_col or not yunmu_col or not diaozhi_col:
    if not shengmu_col or not yunmu_col :
        print(f"错误: 文件 {file_path} 缺少必要的列，跳过处理。")
        return False  # 处理失败

    # 合并“聲母”、“韻母”、“調值”列生成新列"syllable"
    df['syllable'] = df.apply(merge_syllable, axis=1, args=(shengmu_col, yunmu_col, diaozhi_col))
    # df['syllable'] = df.apply(merge_syllable, axis=1, args=(shengmu_col, yunmu_col))
    # 查找“字”列并重命名为“phrase”
    if char_col:
        df['phrase'] = df[char_col]

    # 查找“備註”列，将其中的“文讀”二字清空，并重命名为“notes”
    if comment_col:
        # 清理“備註”列，并处理空值情况
        df['notes'] = df[comment_col].apply(clean_notes)

    # 创建一个新表，只保留"syllable", "phrase", "notes"三列
    columns_to_save = ['syllable', 'phrase', 'notes']
    new_df = df[[col for col in columns_to_save if col in df.columns]]

    # 获取新文件名，基于工作表名中的所有汉字
    new_filename_part = extract_all_chinese_from_sheetname(first_sheet_name)
    new_filename = f"{new_filename_part}.xlsx"
    new_file_path = os.path.join(os.path.dirname(file_path), new_filename)

    # 保存为新的Excel文件
    new_df.to_excel(new_file_path, index=False)

    print(f"处理完成: {new_file_path}")
    return True  # 处理成功


# 主程序
if __name__ == "__main__":
    # 选择文件
    selected_files = select_files()

    # 用于记录处理成功和失败的文件列表
    success_files = []
    failed_files = []

    # 处理每个文件
    for file in selected_files:
        result = process_file(file)
        if result:
            success_files.append(file)
        else:
            failed_files.append(file)

    # 输出处理结果
    print("\n处理结果:")
    print("成功处理的文件:")
    for f in success_files:
        print(f" - {f}")

    if failed_files:
        print("\n未能处理的文件（缺少必要的列）:")
        for f in failed_files:
            print(f" - {f}")
